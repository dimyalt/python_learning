from openpyxl import *
from openpyxl.styles import Border, Side, Font, Alignment

book = load_workbook("C:\\Obrabotka\\Акцизы.xlsx", read_only=True, data_only=True)  # Загружаем исходный файл
sheet = book.active  # Определяем активный лист исходного файла
book.iso_dates = True  # Разрешаем использовать в датах формат ISO 8601
row_count = sheet.max_row - 3  # Переменная, указывающая на количество строк в табличной части

reestr = open("C:\\Obrabotka\\data_file.xlsx")  # Открываем рабочий файл
sheet_r = reestr.active  # Определяем активный лист рабочего файла
reestr.iso_dates = True  # Разрешаем использовать в датах формат ISO 8601

# reestr_row = ['' for _ in range(sheet_r.max_column)] # Создаем массив на все столбцы с пустыми значениями в ячейках

for row in range(7, row_count):  # Начиная с 7й строки и до конца строк в табличной части
    reestr_row2 = ['' for _ in
                   range(sheet_r.max_column)]  # Создаем массив на все столбцы с пустыми значениями в ячейках
    nomer_doc = sheet[row][21].value  # Переменная со значением ячейки из столбца 22 исходной таблицы (Номер купажа)
    reestr_row2.insert(11, nomer_doc)  # Добавляем значение переменной в ячейку 11 столбца (№ документа) рабочего файла
    data_kupaga = sheet[row][20].value  # Переменная со значением ячейки из столбца 21 исходной таблицы (Дата купажа)
    reestr_row2.insert(12, data_kupaga)  # Добавляем значение переменной в ячейку 12 столбца (Дата) рабочего файла
    vsego_litri = sheet[row][19].value  # Переменная со значением ячейки из столбца 20 исходной таблицы (Количество)
    reestr_row2.insert(13,
                       vsego_litri)  # Добавляем значение переменной в ячейку 13 столбца (Всего литров) рабочего файла
    vichet_litri = sheet[row][7].value  # Переменная со значением ячейки из столбца 8 исходной таблицы (Количество)
    reestr_row2.insert(14,
                       vichet_litri)  # Доб. знач. перем. в ячейку 14 столбца (Всего литров подакцизного) рабочего файла
    kontragent_name = sheet[row][24].value  # Переменная со значением ячейки из столбца 25 исходной таблицы (Поставщик)
    reestr_row2.insert(15,
                       kontragent_name)  # Доб. знач. перем. в ячейку 15 столбца (Наим. конрагента) рабочего файла
    kontragent_inn = sheet[row][
        25].value  # Переменная со значением ячейки из столбца 26 исходной таблицы (ИНН поставщика)
    reestr_row2.insert(16, kontragent_inn)  # Добавляем значение переменной в ячейку 16 столбца (ИНН) рабочего файла
    name_doc_2 = sheet[row][
        30].value  # Переменная со значением ячейки из столбца 31 исходной таблицы (Наименование док-та)
    reestr_row2.insert(20,
                       name_doc_2)  # Доб. знач. перем. в ячейку 20 столбца (Наименование док-та) рабочего файла
    nomer_doc_2 = sheet[row][
        29].value  # Переменная со значением ячейки из столбца 30 исходной таблицы (№ док-та поступления)
    reestr_row2.insert(21,
                       nomer_doc_2)  # Доб. знач. перем. в ячейку 21 столбца (№ док-та поступления) рабочего файла
    date_doc_2 = sheet[row][
        28].value  # Переменная со значением ячейки из столбца 29 исходной таблицы (Дата док-та поступления)
    reestr_row2.insert(22,
                       date_doc_2)  # Доб. знач. перем. в ячейку 22 столбца (Дата док-та поступления) рабочего файла
    vsego_podakciz = sheet[row][
        31].value  # Переменная со значением ячейки из столбца 32 исходной таблицы (Кол-во сырья)
    reestr_row2.insert(23,
                       vsego_podakciz)  # Добавляем значение переменной в ячейку 23 столбца (Всего литры) рабочего файла
    vichet_podakciz = sheet[row][
        23].value  # Переменная со значением ячейки из столбца 24 исходной таблицы (Кол-во сырья в купаже)
    reestr_row2.insert(24,
                       vichet_podakciz)  # Доб. знач. перем. в ячейку 24 столбца (Объем подакц. прод.) рабочего файла

    print(reestr_row2)  # Отображаем список
    sheet_r.append(reestr_row2)  # Добавляем список в строку рабочего файла

sheet_r_2 = reestr['Лист2']  # Создаем новую переменную и определяем активный лист рабочего файла
_ = [sheet_r.append(row) for row in sheet_r_2.iter_rows(min_row=1, max_row=2, values_only=True)]  # Копируем строку из
# второго листа в конец табличной части первого

# sheet_r_maxrow = 'AD' + str(sheet_r.max_row)

thin_border = Border(left=Side(style='thin'),  # Создаем переменную с описанием границ
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

for row_ in sheet_r.iter_rows(min_row=23, max_col=30, max_row=sheet_r.max_row - 1):  # Определяем диапазон таблицы
    for cell in row_:  # Для каждой ячейки таблицы
        cell.border = thin_border  # Определяем границы ячейки
        cell.font = Font(size=7)  # Определяем размер шрифта ячейки
        cell.alignment = Alignment(horizontal='center', wrapText=True)  # Определяем центрирование и расположение текста

column_m = sheet_r['M']  # Создаем диапазон на весь столбец М
column_w = sheet_r['W']  # Создаем диапазон на весь столбец М
numbers_string = sheet_r["C22":"AD22"]  # Создаем диапазон строки для нумерации столбцов
for cell in column_m:  # Для всех ячеек в столбце М
    cell.number_format = 'DD.MM.YYYY'  # Устанавливаем формат отображения даты ДД.ММ.ГГГГ

for cell in column_w:  # Для всех ячеек в столбце W
    cell.number_format = 'DD.MM.YYYY'  # Устанавливаем формат отображения даты ДД.ММ.ГГГГ

for row in numbers_string:  # Для всех ячеек диапазона
    counter = 2  # Стартовое значение счетчика (так как первая строка состоит из двух столбцов, я не стал их трогать)
    for cell in row:  # Для всех ячеек в строке
        cell.value = str(counter)  # Вносим значение счетчика в строку
        counter += 1  # Именяем значение счетчика

last_empty_row = len(list(sheet_r.rows))  # Определяет номер последней строки в документе
print(last_empty_row)  # Печатает кол-во строк

_ = [sheet_r.append(row) for row in sheet_r_2.iter_rows(min_row=4, max_row=4, values_only=True)]  # Копируем строку из
# второго листа в конец первого
_ = [sheet_r.append(row) for row in sheet_r_2.iter_rows(min_row=5, max_row=5, values_only=True)]  # Копируем строку из
# второго листа в конец первого
_ = [sheet_r.append(row) for row in sheet_r_2.iter_rows(min_row=6, max_row=6, values_only=True)]  # Копируем строку из
# второго листа в конец первого

reestr.save("C:\\Obrabotka\\Реестр.xlsx")  # Сохраняем как новый файл Реестр.xlsx
