from openpyxl import *
import telebot
import time
from datetime import date
import pandas as pd
from telebot import types

file = load_workbook("sample.xlsx")  # Открываем рабочую книгу / Opening the workbook
list_ = file.active  # Создали переменную для активного листа / Create a variable for the active sheet
file.iso_dates = True  # Сохранение даты и времени в формате ISO 8601 / Store dates and times in the ISO 8601 format
data = []  # Массив, в который добавляется информация от пользователя / The array into which the information from the
# user is added
start_date = ''  # Строка, в которую добавляется информация из сообщения (Дата начала) / String where the information
# from the message is added (Start Date)
end_date = ''  # Строка, в которую добавляется информация из сообщения (Дата окончания) / String where the information
# from the message is added (End Date)
get_name_game_ = ''  # Строка, в которую добавляется информация из сообщения (Название) / String where the information
# from the message is added (Name)
get_org_name = ''  # Строка, в которую добавляется информация из сообщения (Организатор) / String where the information
# from the message is added (Organizer)
get_name_user = ''  # Строка, в которую добавляется информация из сообщения (Имя пользователя) / String where the
# information from the message is added (Username)
id_game = ''  # Строка, в которую генерируется информация (ID мероприятия) / The string in which the information is
# generated (event ID)
get_type_game_ = ''

bot = telebot.TeleBot(
    'XXXXXXXX')  # Подключили токен телеграмм бота / We connected the bot's
# telegram token


@bot.message_handler(commands=['start'])  # Обработка команды "start" / Processing the "start" command
def start_message(message):  # Функция обработки сообщения "start" / Function for processing the "start" message
    markup = types.InlineKeyboardMarkup()  # Создаем клавиатуру для ответа на сообщение "start" / Create a keyboard for
    # answering the "start" message
    button_game = types.InlineKeyboardButton('Следующая игра', callback_data='игра')  # Создаем кнопку клавиатуры
    # "Следующая игра", которая покажет ближайшую игру к текущей дате / Create a keyboard button  "Next game", which
    # will show the nearest game to the current date
    button_games = types.InlineKeyboardButton('Ближайшие игры', callback_data='игры')  # Создаем кнопку клавиатуры
    # "Ближайшие игры", которая покажет ближайшие 5 игр к текущей дате / Create a keyboard button "Nearest games",
    # which will show the nearest 5 games to the current date
    button_info = types.InlineKeyboardButton('Что-бы добавить - напиши "добавить"', callback_data='инфо')  # Кнопка
    # подсказка для добавления мероприятия (сама кнопка не активна) / A tooltip button to add an event (the button
    # itself is not active)
    button_info2 = types.InlineKeyboardButton('Что-бы редактировать - напиши "перенос"', callback_data='инфо')  # Кнопка
    # подсказка для редактирования мероприятия (сама кнопка не активна) / Hint button to edit the event (the button
    # itself is not active)
    markup.row(button_game, button_games)  # Строка с кнопками "Следующая игра" и "Ближайшие игры" / Bar with "Next
    # Game" and "Nearest Games" buttons
    markup.row(button_info)  # Строка с кнопкой подсказкой для добавления мероприятия / A line with a tooltip button
    # to add an event
    markup.row(button_info2)  # Строка с кнопкой подсказкой для редактирования мероприятия / A line with a tooltip
    # button to edit the event
    msg = bot.send_message(message.chat.id,
                           'Короче, ты меня запустил, и я в недотрогу играть не буду: напишешь /help — и я расскажу '
                           'тебе о своём функционале. Заодно проверим насколько понятно мануал написан. Хрен его знает,'
                           ' на кой ляд тебе такое хобби сдалось, но я в чужие дела не лезу, хочешь ездить по играм, '
                           'значит есть за что... ',
                           reply_markup=markup)  # Ответ бота на команду "start" / Bot response to "start" command


@bot.message_handler(commands=['help'])  # Обработка команды "help" / Processing the "help" command
def help_message(message):  # Функция обработки сообщения "help" / Function for processing the "help" message
    bot.send_message(message.chat.id, 'Ну что могу предложить, господа?\n\nНапиши "игра" - получишь информацию про '
                                      'ближайшие событие. Всё проверенно. Все актуально.\n\nНапиши "игры" - расскажу о '
                                      'пяти ближайших событиях. Но помни, мероприятия отменяются или переносятся, а '
                                      'орги иногда не вовремя актуализируют объявления. Так что информация бывает даёт '
                                      'осечки, примерно, пятьдесят на пятьдесят. Ничего не поделаешь - эхо СOVID.\n\n'
                                      'Напишешь "добавить" - получишь возможность внести событие в список. И теперь '
                                      'поподробнее о том как добавлять мероприятие:\n1. Сначала предложит внести дату '
                                      'начала.\n2. Потом укажешь когда конец.\n3. Третьим пунктом укажешь название.\n'
                                      '4. Потом организатора напишешь, сюда же можешь указать ссылку на группу игры в '
                                      'соцсетях.\n5. Пятым пунктом укажешь тип игры, только не выдумывай, укажи из '
                                      'этого списка: (фентези ПРИ, airsoft педулово, airsoft милсим, airsoft ПРИ, NERF '
                                      'ПРИ, пати в дурацком, камерная РИ, городская РИ, кабинетная РИ).\nСписок не '
                                      'маленький, но надежный, никто не запутается.\n\nНадо отредактировать информацию '
                                      'по добавленной тобой игре? Так и напиши - "перенос"\n\n\nАхтунг! Что-то может не '
                                      'работать, или работать не так как мы этого ожидаем. (Извини, время еще не '
                                      'выпрямило руки кодера).')  # Ответ бота на команду "help" / Bot response
    # to "help" command


@bot.message_handler(content_types=['text'])  # Обработка текстовых сообщений / Text message processing
def send_text(message):  # Функция обработки текстовых сообщений / Text message processing function
    if message.text.lower() == 'добавить':  # Если сообщение от пользователя "добавить" / If a message from the user "add"
        global data  # Объявляем глобальную переменную data для изменения ее в функции / Declare the global variable
        # data to change it in the function
        data = []  # data равна пустому списку / data equals empty list
        bot.send_message(message.from_user.id,
                         "Ну что ж, давай добавлять событие. Сам ты его организуешь или это кто-то делает за тебя, "
                         "мы готовы внести его в список. Я такого высокого мнения о наших мероприятиях, что сам их "
                         "иногда посещаю.\n\nСперва напиши дату начала в формате ддммгг (пример: 01012021):")  # Бот
        # запрашивает дату начала мероприятия / The bot asks for the start date of the event
        bot.register_next_step_handler(message, get_start_date)  # следующий шаг – функция get_start_date / The next
        # step is the get_start_date function

    elif message.text.lower() == 'игра':  # Если сообщение от пользователя "игра" / If a message from the user "game"
        current_date = date.today().strftime('%d.%m.%Y')  # Определяем текущую дату в формате дд.мм.гггг / Define the
        # current date in the format dd.mm.yyyy
        current_date_sort = date.today().strftime('%Y-%m-%d')  # Определяем текущую дату в формате гггг-мм-дд
        # (потребуется для сортировки нашей таблицы данных) / Define the current date in the format yyyyy-mm-dd (needed
        # to sort our data table)
        onegame(current_date_sort)  # Запускаем функцию onegame с параметром даты для сортировки / Run the onegame
        # function with the date parameter for sorting (мне кажется это лишний шаг, т.к. следующий шаг делает то же)
        one_game = onegame(current_date_sort)  # Создаем список со значениями функции onegame / Create a list with the
        # values of the onegame function
        bot.send_message(message.from_user.id, f'Сегодня {current_date}\n\U0001F538 \U0001F538 \U0001F538 \U0001F538 \n'
                                               f'Ближайшая игра c {one_game[0]} по {one_game[1]}\n'
                                               f'{one_game[2]} от {one_game[3]}\n'
                                               f'тип игры: {one_game[4]}\n')  # Бот отправляет сообщение с данными из
        # списка onegame / The bot sends a message with data from the onegame list

    elif message.text.lower() == 'игры':  # Если сообщение от пользователя "игры" / If a message from the user "games"
        current_date = date.today().strftime('%d.%m.%Y')  # Определяем текущую дату в формате дд.мм.гггг / Define the
        # current date in the format dd.mm.yyyy
        current_date_sort = date.today().strftime('%Y-%m-%d')  # Определяем текущую дату в формате гггг-мм-дд
        # (потребуется для сортировки нашей таблицы данных) / Define the current date in the format yyyyy-mm-dd (needed
        # to sort our data table)
        games(current_date_sort)  # Запускаем функцию games с параметром даты для сортировки / Run the games
        # function with the date parameter for sorting (мне кажется это лишний шаг, т.к. следующий шаг делает то же)
        many_games = games(current_date_sort)  # Создаем список со значениями функции many_games / Create a list with the
        # values of the many_games function
        bot.send_message(message.from_user.id, f'Сегодня {current_date}\nБлижайшие мероприятия:\n1. С {many_games[0]} '
                                               f'по {many_games[1]}\n'
                                               f'{many_games[2]} от {many_games[3]}\n'
                                               f'тип игры: {many_games[4]}\n\U0001F538 \U0001F538 \U0001F538 \U0001F538 \n'
                                               f'2. С {many_games[5]} по {many_games[6]}\n'
                                               f'{many_games[7]} от {many_games[8]}\n'
                                               f'тип игры: {many_games[9]}\n\U0001F538 \U0001F538 \U0001F538 \U0001F538 \n'
                                               f'3. С {many_games[10]} по {many_games[11]}\n'
                                               f'{many_games[12]} от {many_games[13]}\n'
                                               f'тип игры: {many_games[14]}\n\U0001F538 \U0001F538 \U0001F538 \U0001F538 \n'
                                               f'4. С {many_games[15]} по {many_games[16]}\n'
                                               f'{many_games[17]} от {many_games[18]}\n'
                                               f'тип игры: {many_games[19]}\n\U0001F538 \U0001F538 \U0001F538 \U0001F538 \n'
                                               f'5. С {many_games[20]} по {many_games[21]}\n'
                                               f'{many_games[22]} от {many_games[23]}\n'
                                               f'тип игры: {many_games[24]}\n')  # Бот отправляет сообщение с данными из
        # списка many_games / The bot sends a message with data from the many_games list

    elif message.text.lower() == 'перенос':  # Если сообщение от пользователя "перенос" / If a message from the user "transfer"
        current_date_sort = date.today().strftime('%Y-%m-%d')  # Определяем текущую дату в формате гггг-мм-дд
        # (потребуется для сортировки нашей таблицы данных) / Define the current date in the format yyyyy-mm-dd (needed
        # to sort our data table)
        user_list = porting(message, current_date_sort)  # Создаем вложенный список user_list со значением результата
        # выполнения функции porting / Create a nested user_list with the value of the result of the porting function
        for i in user_list:  # Для каждого элемента во вложенном списке user_list / For each item in the user_list sublist
            bot.send_message(message.from_user.id, f'ID {i[7]}\nС {i[0]} по {i[2]}\nНазвание: {i[3]}\nОрганизатор: '
                                                   f'{i[4]}\nТип: {i[5]}')  # Бот отправляет сообщение с данными из
        # каждого элемента списка user_list / The bot sends a message with the data from each item in user_list
        bot.send_message(message.from_user.id, 'Какую запись редактируем? (Введи ID записи):')  # Бот спрашивает ID
        # записи, которую будет редактировать пользователь / The bot asks for the ID of the record to be edited by the
        # user
        bot.register_next_step_handler(message, edit_buttons)  # Бот переходит к выполнению функции edit_buttons
        # (вывода кнопок редактирования записи) / The bot proceeds to execute the edit_buttons function (outputs the
        # edit buttons of the record)

    else:  # Во всех других случаях / In all other cases
        bot.send_message(message.from_user.id, "Я тебя не понимаю. Напиши /help.")  # Бот сообщает что не понимает
        # введеной команды и просит перейти в раздел справки "/help" / The bot reports that it does not understand the
        # command and asks to go to the help section "/help"


@bot.callback_query_handler(func=lambda call: True)  # Обработка нажатий кнопок / Processing of button presses
def handle(call):  # Функция обработки нажатия кнопок / Button press processing function
    message = str(call.data)  # Переменная message принимает строковое значение нажатой кнопки / The message variable
    # takes the string value of the pressed button
    if message == 'игра':  # Если значение кнопки "игра" / If the value of the "game" button
        current_date = date.today().strftime('%d.%m.%Y')  # Определяем текущую дату в формате дд.мм.гггг / Define the
        # current date in the format dd.mm.yyyy
        current_date_sort = date.today().strftime('%Y-%m-%d')  # Определяем текущую дату в формате гггг-мм-дд
        # (потребуется для сортировки нашей таблицы данных) / Define the current date in the format yyyyy-mm-dd (needed
        # to sort our data table)
        onegame(current_date_sort)  # Запускаем функцию onegame с параметром даты для сортировки / Run the onegame
        # function with the date parameter for sorting (мне кажется это лишний шаг, т.к. следующий шаг делает то же)
        one_game = onegame(current_date_sort)  # Создаем список со значениями функции onegame / Create a list with the
        # values of the onegame function
        bot.send_message(call.message.chat.id,
                         f'Сегодня {current_date}\n\U0001F538 \U0001F538 \U0001F538 \U0001F538 \nБлижайшая игра c '
                         f'{one_game[0]} по {one_game[1]}\n'
                         f'{one_game[2]} от {one_game[3]}\n'
                         f'тип игры: {one_game[4]}\n')  # Бот отправляет сообщение с данными из
        # списка onegame / The bot sends a message with data from the onegame list

    elif message == 'игры':  # Если значение кнопки "игры" / If the value of the "games" button
        current_date = date.today().strftime('%d.%m.%Y')  # Определяем текущую дату в формате дд.мм.гггг / Define the
        # current date in the format dd.mm.yyyy
        current_date_sort = date.today().strftime('%Y-%m-%d')  # Определяем текущую дату в формате гггг-мм-дд
        # (потребуется для сортировки нашей таблицы данных) / Define the current date in the format yyyyy-mm-dd (needed
        # to sort our data table)
        games(current_date_sort)  # Запускаем функцию games с параметром даты для сортировки / Run the games
        # function with the date parameter for sorting (мне кажется это лишний шаг, т.к. следующий шаг делает то же)
        many_games = games(current_date_sort)  # Создаем список со значениями функции many_games / Create a list with the
        # values of the many_games function
        bot.send_message(call.message.chat.id, f'Сегодня {current_date}\nБлижайшие мероприятия:\n1. С {many_games[0]} '
                                               f'по {many_games[1]}\n'
                                               f'{many_games[2]} от {many_games[3]}\n'
                                               f'тип игры: {many_games[4]}\n\U0001F538 \U0001F538 \U0001F538 \U0001F538 \n'
                                               f'2. С {many_games[5]} по {many_games[6]}\n'
                                               f'{many_games[7]} от {many_games[8]}\n'
                                               f'тип игры: {many_games[9]}\n\U0001F538 \U0001F538 \U0001F538 \U0001F538 \n'
                                               f'3. С {many_games[10]} по {many_games[11]}\n'
                                               f'{many_games[12]} от {many_games[13]}\n'
                                               f'тип игры: {many_games[14]}\n\U0001F538 \U0001F538 \U0001F538 \U0001F538 \n'
                                               f'4. С {many_games[15]} по {many_games[16]}\n'
                                               f'{many_games[17]} от {many_games[18]}\n'
                                               f'тип игры: {many_games[19]}\n\U0001F538 \U0001F538 \U0001F538 \U0001F538 \n'
                                               f'5. С {many_games[20]} по {many_games[21]}\n'
                                               f'{many_games[22]} от {many_games[23]}\n'
                                               f'тип игры: {many_games[24]}\n')  # Бот отправляет сообщение с данными из
        # списка many_games / The bot sends a message with data from the many_games list

    elif message == 'название':  # Если значение кнопки "название" / If the value of the "name" button
        result_name_type_game = line_address_detection(id_game, message)  # Создаем переменную result_name_type_game со
        # значением функции line_address_detection (адрес ячейки с хранящимся названием мероприятия) / Create a variable
        # result_name_type_game with the value of the function line_address_detection (address of the cell with the
        # stored event name)
        msg = bot.send_message(call.message.chat.id, f'Меняем название "{list_[result_name_type_game].value}" на:')  #
        # Бот спрашивает новое название мероприятия / The bot asks for a new event name
        bot.register_next_step_handler(msg, correct, result_name_type_game)  # Бот переходит к выполнению функции correct
        # с параметром result_name_type_game / The bot proceeds to execute the correct function with the
        # result_name_type_game parameter

    elif message == 'тип':  # Если значение кнопки "тип" / If the value of the "type" button
        result_name_type_game = line_address_detection(id_game, message)  # Создаем переменную result_name_type_game со
        # значением функции line_address_detection (адрес ячейки) / Create a variable
        # result_name_type_game with the value of the function line_address_detection (address of the cell)
        msg = bot.send_message(call.message.chat.id, f'Меняем тип "{list_[result_name_type_game].value}" на:')  #
        # Бот спрашивает новый тип мероприятия / The bot asks for a new type of event
        bot.register_next_step_handler(msg, correct, result_name_type_game)  # Бот переходит к выполнению функции correct
        # с параметром result_name_type_game / The bot proceeds to execute the correct function with the
        # result_name_type_game parameter

    elif message == 'отмена':  # Если значение кнопки "отмена" / If the value of the "cancel" button
        result_cell_game = line_address_detection(id_game, message)  # Создаем переменную result_cell_game со
        # значением функции line_address_detection (адрес ячейки) / Create a variable
        # result_cell_game with the value of the function line_address_detection (address of the cell)
        result_cancel_game = int(result_cell_game[1:])  # Создаем переменную result_cancel_game в которой хранится номер
        # строки для удаления / Create a variable result_cancel_game that stores the number of the line to delete
        bot.send_message(call.message.chat.id, f'Отмена мероприятия с ID: "{list_[result_cell_game].value}"')  # Бот
        # сообщает ID мероприятия, которое будет удалено / The bot reports the ID of the event that will be deleted
        list_.delete_rows(result_cancel_game)  # Удаляем строку под номером, который передала переменная
        # result_cancel_game / Delete the line under the number passed by the variable result_cancel_game
        file.save("sample.xlsx")  # Сохраняем таблицу / Save the table
        sorting()  # Запускаем функцию сортировки записей по дате / Start the function of sorting records by date
        bot.send_message(call.message.chat.id, f'Удалено.')  # Бот сообщает об удалении / The bot reports deletion

    elif message == 'даты':   # Если значение кнопки "даты" / If the value of the "dates" button
        result_start_date_game = line_address_detection(id_game, message)  # Создаем переменную result_start_date_game со
        # значением функции line_address_detection (адрес ячейки) / Create a variable result_start_date_game with the
        # value of the function line_address_detection (cell address)
        msg = bot.send_message(call.message.chat.id, f'Меняем дату начала "{list_[result_start_date_game].value}" на:')
        # Бот спрашивает новую дату начала мероприятия / The bot asks for a new start date for the event
        bot.register_next_step_handler(msg, correct_start_date, result_start_date_game)  # Бот переходит к выполнению
        # функции correct_start_date с параметром result_start_date_game / The bot proceeds to execute the
        # correct_start_date function with the result_start_date_game parameter


print('start')


def date_check(date):  # Функция проверки правильности написания даты при вводе нового мероприятия / Function to check
    # the correct spelling of the date when entering a new event
    try:
        date = time.strptime(date, '%d%m%Y')  # Если дата заполнена в формате ддммгггг / If the date is filled in the
        # format ddmmyyyy
        true_date = time.strftime('%d.%m.%Y',
                                  date)  # Создаем переменную даты в формате дд.мм.гггг (для вывода пользователю) /
        # Create a date variable in the format dd.mm.yyyyy (for output to the user)
        sort_date = time.strftime('%Y-%m-%d',
                                  date)  # Создаем переменную даты в формате гггг-мм-дд (для функции сортировки) /
        # Create a date variable in the format yyyy-mm-dd (for the sorting function)
        return true_date, sort_date  # Возвращаем даты / Returning dates
    except ValueError:  # Если дата не заполнена в формате ддммгггг появляется ошибка / If the date is not filled in
        # the format ddmmyyyy, an error appears
        return False  # Возвращаем False / Return False


def get_start_date(message):  # Функция получения даты начала мероприятия / Function for getting the start date of the
    # event
    global start_date  # Объявляем глобальную переменную start_date для изменения ее в функции / Declare the global
    # variable start_date to change it in the function
    start_date = message.text  # Создаем переменную start_date со значением текста сообщения / Create a variable
    # start_date with the value of the message text
    if not date_check(start_date):  # Если функция проверки даты date_check передала False / If the date_check
        # function passes False
        bot.send_message(message.from_user.id, 'Неправильный формат даты! Попробуй еще раз (например 01012021):')  # Бот
        # отправляет сообщение о неверной дате / The bot sends a message about the wrong date
        bot.register_next_step_handler(message, get_start_date)  # Бот переходит к выполнению функции
        # get_start_date ввода даты снова / The bot proceeds to perform the get_start_date function of entering the
        # date again
    else:  # Если функция проверки даты date_check не передала False / If the date_check function does not pass False
        user_date, sorting_date = date_check(start_date)  # В переменные user_date и sorting_date передаются значения
        # из функции date_check (дата для человека и для сортировки) / The user_date and sorting_date variables receive
        # values from the date_check function (date for the person and for sorting)
        data.append(user_date)  # В список data добавляется значение даты для показа пользователю / A date value is
        # added to the data list to show to the user
        data.append(sorting_date)  # В список data добавляется значение даты для сортировки / A date value is added to
        # the data list for sorting
        bot.send_message(message.from_user.id, 'Теперь не забудь указать когда народ должен разъехаться в том же '
                                               'формате. (пример: 01012021):')  # Бот отправляет сообщение с вопросом о
        # дате окончания мероприятия / The bot sends a message asking about the end date of the event
        bot.register_next_step_handler(message, get_end_date)  # Бот переходит к выполнению функции get_end_date / The
        # bot proceeds to the get_end_date function


def correct_start_date(message, result_start_date_game):  # Функция изменения даты начала мероприятия / Function for
    # changing the start date of an event
    correct_startdate = message.text  # Создаем переменную correct_startdate со значением текста сообщения / Create a
    # variable correct_startdate with the value of the message text
    if not date_check(correct_startdate):  # Если функция проверки даты date_check передала False / If the
        # date_check function passes False
        msg = bot.send_message(message.from_user.id, 'Неправильный формат даты! Попробуй еще раз (например 01012021):')
        # Бот отправляет сообщение о неверной дате / The bot sends a message about the wrong date
        bot.register_next_step_handler(msg, correct_start_date, result_start_date_game)  # Бот переходит к выполнению
        # функции correct_start_date ввода даты снова / The bot proceeds to perform the correct_start_date function of
        # entering the date again
    else:  # Если функция проверки даты date_check не передала False / If the date_check function does not pass False
        correct_user_date, correct_sorting_date = date_check(correct_startdate)  # В переменные correct_user_date и
        # correct_sorting_date передаются значения из функции date_check (дата для человека и для сортировки) /
        # The correct_user_date and correct_sorting_date variables receive values from the date_check function (date
        # for the person and for sorting)
        list_[result_start_date_game] = correct_user_date  # В ячейку с адресом переданным из result_start_date_game
        # записываем значение переменной correct_user_date / In the cell with the address passed from
        # result_start_date_game write the value of variable correct_user_date
        result_correct_sorting_date = 'B' + result_start_date_game[1:]  # Получаем адрес ячейки в которой хранится дата
        # в формате гггг-мм-дд и записываем ее в переменную result_correct_sorting_date / Get the address of the cell
        # where the date is stored in the format yyyyy-mm-dd and write it into the variable result_correct_sorting_date
        list_[result_correct_sorting_date] = correct_sorting_date  # В ячейку с адресом переданным из
        # result_correct_sorting_date записываем значение переменной correct_sorting_date / In the cell with the
        # address passed from result_correct_sorting_date write the value of variable correct_sorting_date
        result_correct_end_date = 'C' + result_start_date_game[1:]  # Получаем адрес ячейки в которой хранится дата
        # окончания мероприятия в формате дд.мм.гггг и записываем ее в переменную result_correct_end_date / Get the
        # address of the cell where the end date of the event is stored in the format dd.mm.yyyyy and record it in the
        # variable result_correct_end_date
        file.save("sample.xlsx")  # Сохраняем таблицу / Save the table
        sorting()  # Запускаем функцию сортировки записей по дате / Start the function of sorting records by date
        bot.send_message(message.from_user.id, f'Ok, записал.')  # Бот сообщает о сохранении / The bot reports on saving
        msg = bot.send_message(message.chat.id, f'Меняем дату окончания "{list_[result_correct_end_date].value}" на:')
        # Бот спрашивает новую дату окончания мероприятия / The bot asks for a new end date of the event
        bot.register_next_step_handler(msg, correct_end_date, result_correct_end_date)  # Бот переходит к выполнению
        # функции correct_end_date с параметром result_correct_end_date / The bot proceeds to the correct_end_date
        # function with the result_correct_end_date parameter


def correct_end_date(message, result_correct_end_date):  # Функция изменения даты конца мероприятия / Function for
    # changing the end date of the event
    correct_enddate = message.text  # Создаем переменную correct_enddate со значением текста сообщения / Create a
    # correct_enddate variable with the value of the message text
    if not date_check(correct_enddate):  # Если функция проверки даты date_check передала False / If the
        # date_check function passes False
        msg = bot.send_message(message.from_user.id, 'Неправильный формат даты! Попробуй еще раз (например 01012021):')
        # Бот отправляет сообщение о неверной дате / The bot sends a message about the wrong date
        bot.register_next_step_handler(msg, correct_end_date, result_correct_end_date)  # Бот переходит к выполнению
        # функции correct_end_date ввода даты снова / The bot proceeds to perform the correct_end_date function of
        # entering the date again
    else:  # Если функция проверки даты date_check не передала False / If the date_check function does not pass False
        correct_user_enddate, correct_sorting_enddate = date_check(correct_enddate)  # В переменные correct_user_enddate
        # и correct_sorting_enddate передаются значения из функции date_check (дата для человека и для сортировки)
        list_[result_correct_end_date] = correct_user_enddate  # В ячейку с адресом переданным из
        # result_correct_end_date записываем значение переменной correct_user_enddate / In the cell with the
        # address passed from result_correct_end_date write the value of variable correct_user_enddate
        file.save("sample.xlsx")  # Сохраняем таблицу / Save the table
        sorting()  # Запускаем функцию сортировки записей по дате / Start the function of sorting records by date
        bot.send_message(message.from_user.id, f'Ok, дата окончания изменена.')  # Бот сообщает, что дата окончания
        # мероприятия изменена / The bot reports that the end date of the event has been changed


def get_end_date(message):  # Функция получения даты окончания мероприятия / Function for getting the end date of the
    # event
    global end_date  # Объявляем глобальную переменную end_date для изменения ее в функции / Declare the global
    # variable end_date to change it in the function
    end_date = message.text  # Создаем переменную end_date со значением текста сообщения / Create a variable
    # end_date with the value of the message text
    if not date_check(end_date):  # Если функция проверки даты date_check передала False / If the date_check
        # function passes False
        bot.send_message(message.from_user.id, 'Неправильный формат даты! Попробуй еще раз (например 01012021):')  # Бот
        # отправляет сообщение о неверной дате / The bot sends a message about the wrong date
        bot.register_next_step_handler(message, get_end_date)  # Бот переходит к выполнению функции
        # get_end_date ввода даты снова / The bot proceeds to perform the get_end_date function of entering the
        # date again
    else:  # Если функция проверки даты date_check не передала False / If the date_check function does not pass False
        user_date, sorting_date = date_check(end_date)  # В переменные user_date и sorting_date передаются значения
        # из функции date_check (дата для человека и для сортировки) / The user_date and sorting_date variables receive
        # values from the date_check function (date for the person and for sorting)
        data.append(user_date)  # В список data добавляется значение даты для показа пользователю / A date value is
        # added to the data list to show to the user
        bot.send_message(message.from_user.id, 'Так, ну и как называется вся эта содомия?')  # Бот отправляет сообщение
        # с вопросом о названии мероприятия / The bot sends a message with a question about the name of the event
        bot.register_next_step_handler(message, get_name_game)  # Бот переходит к выполнению функции get_name_game / The
        # bot proceeds to the get_name_game function


def get_name_game(message):  # Функция получения даты названия мероприятия / Function for getting the start name of the
    # event
    global get_name_game_  # Объявляем глобальную переменную get_name_game для изменения ее в функции / Declare the global
    # variable get_name_game to change it in the function

    get_name_game_ = message.text  # Создаем переменную get_name_game со значением текста сообщения / Create a variable
    # get_name_game with the value of the message text
    data.append(get_name_game_)  # В список data добавляется значение текста сообщения / The message text value is added
    # to the data list

    bot.send_message(message.from_user.id, 'Сейчас организатора пиши. Группу там мастерскую али организацию какую. '
                                           'Общественность должна знать кого в случае чего поливать гуано. Ну и указывай'
                                           ' ссылку на группу игры в соцсетях.')  # Бот отправляет сообщение с вопросом о
    # организаторе мероприятия / The bot sends a message with a question about the event organizer
    bot.register_next_step_handler(message, get_orgname)  # Бот переходит к выполнению функции get_orgname / The
    # bot proceeds to the get_orgname function


def get_orgname(message):  # Функция получения организатора мероприятия / The function of getting an event organizer
    global get_orgname  # Объявляем глобальную переменную get_orgname для изменения ее в функции / Declare the global
    # variable get_orgname to change it in the function
    get_org_name = message.text  # Создаем переменную get_org_name со значением текста сообщения / Create a variable
    # get_org_name with the value of the message text
    data.append(get_org_name)  # В список data добавляется значение текста сообщения / The message text value is added
    # to the data list
    bot.send_message(message.from_user.id, 'Ну и финал - давай определимся с типом игры, только не выдумывай, укажи из '
                                           'этого списка: (фентези ПРИ, airsoft педулово, airsoft милсим, airsoft ПРИ, '
                                           'NERF ПРИ, пати в дурацком, камерная РИ, городская РИ, кабинетная РИ).')  #
    # Бот отправляет сообщение с вопросом о типе игры / The bot sends a message with a question about the type of game
    bot.register_next_step_handler(message, get_type_game)  # Бот переходит к выполнению функции get_type_game / The
    # bot proceeds to the get_type_game function
    print(data)


def get_type_game(message):  # Функция получения типа мероприятия и имени пользователя телеграм, отправившего сообщение
    # / Function for getting the event type and the name of the telegram user who sent the message
    global get_type_game_  # Объявляем глобальную переменную get_type_game для изменения ее в функции / Declare the
    # global variable get_type_game to change it in the function
    global get_name_user  # Объявляем глобальную переменную get_name_user для изменения ее в функции / Declare the
    # global variable get_name_user to change it in the function
    get_type_game_ = message.text  # Создаем переменную get_type_game со значением текста сообщения / Create a variable
    # get_type_game with the value of the message text
    get_name_user = message.from_user.username  # Создаем переменную get_name_user со значением автора сообщения /
    # Create a get_name_user variable with the value of the author of the message
    data.append(get_type_game_)  # В список data добавляется значение переменной get_type_game / The value of the
    # get_type_game variable is added to the data list
    data.append(get_name_user)  # В список data добавляется значение переменной get_name_user / The value of the
    # get_name_user variable is added to the data list
    get_id_game()  # Вызываем функцию создания ID записи / Call the record ID creation function
    list_.append(data)  # В активный лист документа добавляем значения списка data / In the active sheet of the
    # document we add the values of the data list
    file.save("sample.xlsx")  # Сохраняем документ / Save the document
    sorting()  # Вызываем функцию создания сортировки записей по дате / Call the function to create a sorting of
    # records by date
    bot.send_message(message.from_user.id, f'Всё, справились. Вот тебе \U0001F194 твоей записи - {data[-1]}, на случай '
                                           f'если надо будет что-то изменить.')  # Бот отправляет сообщение что
    # мероприятие записано / The bot sends a message that the event is recorded


def get_id_game():  # Функция создания ID записи / Record ID creation function
    global get_id_game  # Объявляем глобальную переменную get_id_game для изменения ее в функции / Declare the
    # global variable get_id_game to change it in the function
    id_cell = 1  # Переменная, в которую запишем максимальное значение ID / The variable where we write the maximum
    # ID value
    for row in range(2, list_.max_row + 1):  # Для строк со второй и до последней / For lines from the second to the
        # last
        for column in "H":  # Для столбца H в котором указан ID / For column H, which contains the ID
            cell_name = "{}{}".format(column, row)  # Получаем адрес ячейки / Get the cell address
            if list_[cell_name].value > id_cell:  # Если значение ячейки больше значения id_cell / If the value of the
                # cell is greater than id_cell
                id_cell = list_[cell_name].value  # Присваеваем значение ячейки переменной id_cell / Assign a cell
                # value to the id_cell variable
    data.append(id_cell + 1)  # В список data добавляется значение переменной id_cell + 1 / The value of the id_cell
    # variable + 1 is added to the data list


def onegame(current_date_sort):  # Функция показа ближайшего мероприятия / Function to show the next event
    temp = []  # Временный список / Temporary list
    result = []  # Результат / Result
    list_row = []  # Список со значениями номера ячеек, даты которых больше текущей даты / A list with the values of
    # the cell numbers whose dates are larger than the current date
    search_date = current_date_sort  # Дата поиска = переданному значению / Search Date = Passed Value

    for row in range(2, list_.max_row + 1):  # Для строк со второй и до последней / For lines from the second to the
        # last
        for column in "B":  # Для столбца B в котором указана дата в формате для поиска ГГГГ-мм-дд / For column B,
            # which contains the date in search format yyyy-mm-dd
            cell_name = "{}{}".format(column, row)  # Получаем адрес ячейки / Get the cell address
            if list_[cell_name].value >= search_date:  # Если значение ячейки больше или равно текущей дате / If the
                # cell value is greater than or equal to the current date
                list_row.append(row)  # Значение ячейки добавляем в список / The cell value is added to the list

    for row in list_.iter_rows(min_row=list_row[0], max_col=6, max_row=list_row[0]):  # Для диапазона начиная со строки
        # № которой идет 0 в списке дат и на протяжении всех столбцов / For the range starting with row № which goes 0
        # in the date list and throughout all columns
        for cell in row:  # Для всех ячеек / For all cells
            temp.append(cell.value)  # Добавляем во временный список значение ячеек / Adding cell values to a temporary
            # list

    result.append(temp[0])  # Добавляем в итоговый список значение начала / Add the start value to the summary list
    result.append(temp[2])  # Добавляем в итоговый список значение конца / Add an end value to the final list
    result.append(temp[3])  # Добавляем в итоговый список значение названия / Add the value of the name to the final
    # list
    result.append(temp[4])  # Добавляем в итоговый список значение организатора / Add the value of the organizer to the
    # final list
    result.append(temp[5])  # Добавляем в итоговый список значение типа / Add a value of type to the resulting list

    return result  # Возвращаем result / Return result


def sorting():  # Функция сортировки записей в таблице мероприятий / Function for sorting records in the table of events
    xl = pd.ExcelFile("sample.xlsx")  # Открываем pandas наш файл / Open our pandas file
    df = xl.parse("testexel")  # Открываем pandas активный лист / Open pandas active sheet
    df = df.sort_values(by="Сортировка")  # Сортируем по возрастанию столбец Сортировка (sort_date ('%Y-%m-%d')) / Sort
    # in ascending order column Sort (sort_date ('%Y-%m-%d'))
    writer = pd.ExcelWriter('sample.xlsx')  # Записываем файл / Writing down the file
    df.to_excel(writer, sheet_name='testexel', columns=["Начало", "Сортировка", "Конец", "Название", "Организатор",
                                                        "Тип", "Владелец", "ID"],
                index=False)  # Выбираем какие столбцы записывать / Choose which columns to write
    writer.save()  # Сохраняем / Save file
    # Просмотр таблицы / View table
    for i in range(0, list_.max_row):  # Для всех строк от 0 до максимальной / For all lines from 0 to the maximum
        for col in list_.iter_cols(1, list_.max_column):  # Для всех колонок от 0 до максимальной / For all columns
            # from 0 to maximum
            print(col[i].value, end="\t\t")  # Печатаем значение ячейки / Print the cell value
        print('')


def porting(message, current_date_sort):  # Функция редактирования записей / Record editing function
    global get_name_user  # Объявляем глобальную переменную get_name_user для изменения ее в функции / Declare the
    # global variable get_name_user to change it in the function
    search_date = current_date_sort  # Дата поиска = переданному значению / Search Date = Passed Value
    temp_name = []  # Временный список со значением названия мероприятия / Temporary list with event name value
    temp_date = []  # Временный список со значением даты мероприятия / Temporary list with event date value
    temp_result = []  # Временный список с результатом / Temporary list with result
    temp = []
    get_name_user = message.from_user.username  # Получаем имя пользователя, написавшего сообщение / Get the name of
    # the user who wrote the message
    for row in range(2, list_.max_row + 1):  # Для строк со второй и до последней / For lines from the second to the
        # last
        for column in "G":  # Для столбца G в котором указано имя пользователя, создавшего запись / For column G, which
            # contains the name of the user who created the record
            cell_name = "{}{}".format(column, row)  # Получаем адрес ячейки / Get the cell address
            if list_[cell_name].value == get_name_user:  # Если значение ячейки равно имени пользователя / If the cell
                # value is equal to the user name
                temp_name.append(row)  # Значение ячейки добавляем в список / The cell value is added to the list

    for row in range(2, list_.max_row + 1):  # Для строк со второй и до последней / For lines from the second to the
        # last
        for column in "B":  # Для столбца B в котором указана дата в формате для поиска ГГГГ-мм-дд / For column B,
            # which contains the date in search format yyyy-mm-dd
            cell_name = "{}{}".format(column, row)  # Получаем адрес ячейки / Get the cell address
            if list_[cell_name].value >= search_date:  # Если значение ячейки больше или равно текущей дате / If the
                # cell value is greater than or equal to the current date
                temp_date.append(row)  # Значение ячейки добавляем в список / The cell value is added to the list

    for i in temp_name:  # Для каждого значения в списке temp_name
        if i in temp_date:  # Для каждого значения в списке temp_date
            temp_result.append(i)  # Итоговое значение ячейки добавляем в список temp_result / The final value of the
            # cell is added to the list temp_result
    print(temp_result)

    for row in list_.iter_rows(min_row=temp_result[0], max_col=8, max_row=list_.max_row):  # Для диапазона начиная
    # со строки № которой идет 0 в списке и на протяжении всех столбцов / For the range starting with row # which
    # goes 0 in the list and throughout all columns
        for cell in row:  # Для всех ячеек / For all cells
            temp.append(cell.value)  # Добавляем во временный список значение ячеек / Adding cell values to a temporary
            # list

    sep = 8  # кол-во элементов в одном внутреннем списке / Number of items in one internal list
    result = [temp[x:x + sep] for x in range(0, len(temp), sep)]  # Разбиваем список temp на части по 8 элементов и
    # добавляем их как вложенные списки в result / Split the temp list into parts of 8 elements and add them as
    # sublists in result

    return result  # Возвращаем список result / Return the list result


def edit_buttons(message):  # Функция вывода кнопок редактирования записи / Record edit buttons output function
    global id_game  # Объявляем глобальную переменную id_game для изменения ее в функции / Declare the
    # global variable id_game to change it in the function
    id_game = message.text  # В месседже получаем ID игры, которую будем редактировать / In the message we get the ID of
    # the game we are going to edit
    markup = types.InlineKeyboardMarkup()  # Разметка клавиатуры / Keyboard markup
    button_startend_game = types.InlineKeyboardButton('Даты проведения', callback_data='даты')  # Создаем кнопку Даты
    # проведения / Create a Date button
    button_name_game = types.InlineKeyboardButton('Название', callback_data='название')  # Создаем кнопку Название /
    # Create a Name button
    button_type_game = types.InlineKeyboardButton('Тип', callback_data='тип')  # Создаем кнопку Тип / Create a Type
    # button
    button_cancel_game = types.InlineKeyboardButton('Отмена мероприятия', callback_data='отмена')  # Создаем кнопку
    # Отмена / Create a Cancel button
    markup.row(button_startend_game, button_name_game)  # Создаем строку с кнопками "Даты" и "Название"
    # / Create a line with the "Dates" and "Name" buttons
    markup.row(button_type_game, button_cancel_game)  # Создаем строку с кнопками "Тип" и "Отмена" / Create a line with
    # the "Type" and "Cancel" buttons
    bot.send_message(message.chat.id, 'Выбери что редактируем?', reply_markup=markup)  # Бот отправляет сообщение
    # с просьбой сделать выбор и нажать на кнопку / The bot sends a message with a request to make a choice and click


def line_address_detection(id_game, message):  # Функция определения адреса ячейки / Function for determining the cell
    # address
    for row in range(2, list_.max_row + 1):  # Для строк со второй и до последней / For lines from the second to the
        # last
        for column in "H":  # Для столбца H в котором указано ID мероприятия / For column H, which contains the event ID
            cell_name = "{}{}".format(column, row)  # Получаем адрес ячейки / Get the cell address
            string_num = "{}".format(row)  # Получаем номер строки / Get the line number
            if str(list_[cell_name].value) == id_game:  # Если значение в ячейке совпадает с введенным ID / If the
                # value in the cell matches the entered ID
                if message == 'название':  # Если в сообщениие было указано "название" / If "title" was specified in
                    # the message
                    name_cell_adr = 'D' + string_num  # Создаем переменную с адресом ячейки с названием мероприятия /
                    # Create a variable with the address of the cell with the name of the event
                elif message == 'тип':  # Если в сообщениие было указано "тип" / If "type" was specified in
                    # the message
                    name_cell_adr = 'F' + string_num  # Создаем переменную с адресом ячейки с типом мероприятия /
                    # Create a variable with the address of a cell with the type of event
                elif message == 'отмена':  # Если в сообщениие было указано "отмена" / If "cancel" was specified in
                    # the message
                    name_cell_adr = 'H' + string_num  # Создаем переменную с адресом ячейки с ID мероприятия / Create
                    # a variable with the address of the event ID cell
                elif message == 'даты':  # Если в сообщениие было указано "даты" / If "dates" was specified in
                    # the message
                    name_cell_adr = 'A' + string_num  # Создаем переменную с адресом ячейки с началом мероприятия /
                    # Create a variable with the address of the cell with the beginning of the event
                return name_cell_adr  # Возвращает адрес ячейки / Returns the cell address


def correct(message, result_name_type_game):  # Функция внесения изменений в записи файла / Function for making changes
    # to file entries
    list_[result_name_type_game] = message.text  # В ячейку с переданным адресом вносим изменения из текста сообщения /
    # In the cell with the passed address we make changes from the text of the message
    file.save("sample.xlsx")  # Сохраняем таблицу / Save the table
    sorting()  # Запускаем функцию сортировки записей по дате / Start the function of sorting records by date
    bot.send_message(message.from_user.id, f'Ok, записал.')  # Бот сообщает о сохранении / The bot reports on saving


def games(current_date_sort):
    temp = []  # Временный массив
    result = []  # Результат
    list_row = []  # Список со значениями номера ячеек, даты которых больше текущей даты
    search_date = current_date_sort  # Дата поиска = переданному значению

    for row in range(2, list_.max_row + 1):  # Для строк со второй и до последней
        for column in "B":  # Для столбца B в котором указана дата в формате для поиска ГГГГ-мм-дд
            cell_name = "{}{}".format(column, row)  # Получаем адрес ячейки
            if list_[cell_name].value >= search_date:  # Если значение ячейки больше или равно текущей дате
                list_row.append(row)  # Значение ячейки добавляем в список

    # for row in list_.iter_rows(min_row=list_row[0], max_col=6, max_row=list_row[4]):
    for row in list_.iter_rows(min_row=list_row[0], max_col=6,
                               max_row=list_.max_row):  # Для диапазона начиная со строки
        # № которой идет 0 в списке дат и на протяжении всех столбцов
        for cell in row:  # Для всех ячеек
            temp.append(cell.value)  # Добавляем во временный список значение ячеек
    # print(len(temp))

    if len(temp) < 30:
        while len(temp) < 30:
            temp.append('-')
    # print(temp)

    result.append(temp[0])  # Добавляем в итоговый список значение начала
    result.append(temp[2])  # Добавляем в итоговый список значение конца
    result.append(temp[3])  # Добавляем в итоговый список значение названия
    result.append(temp[4])  # Добавляем в итоговый список значение организатора
    result.append(temp[5])  # Добавляем в итоговый список значение типа

    result.append(temp[6])  # Добавляем в итоговый список значение начала второй игры
    result.append(temp[8])  # Добавляем в итоговый список значение конца второй игры
    result.append(temp[9])  # Добавляем в итоговый список значение названия второй игры
    result.append(temp[10])  # Добавляем в итоговый список значение организатора второй игры
    result.append(temp[11])  # Добавляем в итоговый список значение типа второй игры

    result.append(temp[12])  # Добавляем в итоговый список значение начала третьей игры
    result.append(temp[14])  # Добавляем в итоговый список значение конца третьей игры
    result.append(temp[15])  # Добавляем в итоговый список значение названия третьей игры
    result.append(temp[16])  # Добавляем в итоговый список значение организатора третьей игры
    result.append(temp[17])  # Добавляем в итоговый список значение типа третьей игры

    result.append(temp[18])  # Добавляем в итоговый список значение начала четвертой игры
    result.append(temp[20])  # Добавляем в итоговый список значение конца четвертой игры
    result.append(temp[21])  # Добавляем в итоговый список значение названия четвертой игры
    result.append(temp[22])  # Добавляем в итоговый список значение организатора четвертой игры
    result.append(temp[23])  # Добавляем в итоговый список значение типа четвертой игры

    result.append(temp[24])  # Добавляем в итоговый список значение начала пятой игры
    result.append(temp[26])  # Добавляем в итоговый список значение конца пятой игры
    result.append(temp[27])  # Добавляем в итоговый список значение названия пятой игры
    result.append(temp[28])  # Добавляем в итоговый список значение организатора пятой игры
    result.append(temp[29])  # Добавляем в итоговый список значение типа пятой игры

    return result


bot.polling(none_stop=True,
            interval=0)  # Опрос телеграмма на входящие сообщения / Telegram survey for incoming messages
